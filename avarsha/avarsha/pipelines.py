# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html



import rfc822
import time
from cStringIO import StringIO
from PIL import Image
from sets import Set
from twisted.internet import defer

from scrapy import log
from scrapy.contrib.pipeline.files import S3FilesStore
from scrapy.contrib.pipeline.images import ImagesPipeline
from scrapy.exceptions import DropItem
from boto.dynamodb2.exceptions import ItemNotFound

from avarsha.dynamo_db import DynamoDB
from openpyxl import load_workbook


class AvarshaPipeline(object):
    def __init__(self):
        self.dynamo_db = DynamoDB()

    def __build_product_availabity_item(self, product_item):
        if (product_item.get('colors') is None
            and product_item.get('sizes') is None
            and product_item.get('stocks') is None):
            return None

        pa_item = {}
        pa_item['product_id'] = product_item['product_id']
        if product_item.get('colors') != None:
            pa_item['colors'] = product_item['colors']
        if product_item.get('sizes') != None:
            pa_item['sizes'] = product_item['sizes']
        if product_item.get('stocks') != None:
            pa_item['stocks'] = product_item['stocks']
        pa_item['crawl_datetime'] = product_item['crawl_datetime']
        return pa_item

    def __build_product_sales_item(self, product_item):
        if (product_item.get('list_price') is None
            and product_item.get('price') is None
            and product_item.get('low_price') is None
            and product_item.get('high_price') is None
            and product_item.get('free_shipping') is None):
            return None

        ps_item = {}
        ps_item['product_id'] = product_item['product_id']
        if product_item.get('list_price') != None:
            ps_item['list_price'] = product_item['list_price']
        if product_item.get('price') != None:
            ps_item['price'] = product_item['price']
        if product_item.get('low_price') != None:
            ps_item['low_price'] = product_item['low_price']
        if product_item.get('high_price') != None:
            ps_item['high_price'] = product_item['high_price']
        if product_item.get('free_shipping') != None:
            ps_item['free_shipping'] = product_item['free_shipping']
        ps_item['crawl_datetime'] = product_item['crawl_datetime']
        return ps_item

    def __build_product_reviews_item(self, product_item):
        if (product_item.get('review_count') is None
            and product_item.get('max_review_rating') is None
            and product_item.get('review_rating') is None
            and product_item.get('review_list') is None):
            return None

        pr_item = {}
        pr_item['product_id'] = product_item['product_id']
        if product_item.get('review_count') != None:
            pr_item['review_count'] = product_item['review_count']
        if product_item.get('max_review_rating') != None:
            pr_item['max_review_rating'] = product_item['max_review_rating']
        if product_item.get('review_rating') != None:
            pr_item['review_rating'] = product_item['review_rating']
        if product_item.get('review_list') != None:
            pr_item['review_list'] = product_item['review_list']
        pr_item['crawl_datetime'] = product_item['crawl_datetime']
        return pr_item

    def process_item(self, item, spider):
        if item.normalize_attributes() is False:
            raise DropItem("Attributes format error in %s" % item)
        #self.__assert_necessary_attributes(item)

#         if spider.settings['VERSION'] == 'DEV':
#             return item
        if spider.settings['VERSION'] == 'DEV':
            #self.store_to_excel(item)
            self.init_to_excel(item)
            return item

        if spider.settings['CHROME_ENABLED'] is True:
            try:
                chrome_item = self.dynamo_db.table('chrome').get_item(url=item['url'])
                item['chrome_clicks'] = chrome_item['clicks']
            except ItemNotFound:
                log.msg('URL not in chrome log: %s' % item['url'] , log.DEBUG)

        product_id = item.get('product_id')
        old_item = None
        try:
            old_item = self.dynamo_db.table('products').get_item(product_id=product_id)
            item['collections'] = list(item['collections'])
            if old_item != item:
                item['updated_datetime'] = item['crawl_datetime']
                log.msg('Item updated since last crawl: %s' % product_id, log.DEBUG)
            else:
                # detect if it is crawled recently
                old_item['crawl_datetime'] = item['crawl_datetime']
                old_item.partial_save()
                log.msg('Item not changed: %s' % product_id, log.DEBUG)
                return item
        except ItemNotFound:
            log.msg('Find a new item: %s' % product_id, log.DEBUG)
            item['updated_datetime'] = item['crawl_datetime']

        # merge current collections to previous collections
        if old_item != None:
            new_collections = Set(old_item.get('collections')) \
                | Set(item.get('collections'))
        else:
            new_collections = Set(item.get('collections'))
        if len(new_collections) != 0:
            item['collections'] = list(new_collections)

        log.msg('Save item info to Dynamo DB: %s' % product_id, log.DEBUG)

        self.dynamo_db.table('products').put_item(data=item, overwrite=True)

        pa_item = self.__build_product_availabity_item(item)
        if pa_item != None:
            self.dynamo_db.table('product_availabity')\
                .put_item(data=pa_item, overwrite=True)

        ps_item = self.__build_product_sales_item(item)
        if ps_item != None:
            self.dynamo_db.table('product_sales')\
                .put_item(data=ps_item, overwrite=True)

        pr_item = self.__build_product_reviews_item(item)
        if pr_item != None:
            self.dynamo_db.table('product_reviews')\
                .put_item(data=pr_item, overwrite=True)

        return item

    def open_spider(self, spider):
        feeder = spider.feeder

        if spider.settings['VERSION'] == 'DEV':
            # for test for spider
            start_urls = ['http://www.sophiaprom.com/special-occasion-dresses-catid673004416',
                          'http://www.sophiaprom.com/evening-dresses-catid673004420',
                          'http://www.sophiaprom.com/ball-gown-dresses-catid673007988',
                          'http://www.sophiaprom.com/cocktail-dresses-catid673004419',
                          'http://www.sophiaprom.com/new-arrival-catid673004423',
                          'http://www.sophiaprom.com/prom-dresses-2016-catid673004421',
                          'http://www.sophiaprom.com/two-piece-prom-dresses-catid786017081',
                          'http://www.sophiaprom.com/quinceanera-dresses-catid673004441',
                          'http://www.sophiaprom.com/sweet-16-catid673007840',
                          'http://www.sophiaprom.com/homecoming-dresses-catid673004436',
                          'http://www.sophiaprom.com/formal-dresses-catid673005550',
                          'http://www.sophiaprom.com/wedding-dresses-catid673000142',
                          'http://www.sophiaprom.com/aline-wedding-dresses-catid673005546',
                          'http://www.sophiaprom.com/2016-new-arrival-catid673004349',
                          'http://www.sophiaprom.com/beach-wedding-dresses-catid673003635',
                          'http://www.sophiaprom.com/casual-wedding-dress-catid673005544',
                          'http://www.sophiaprom.com/plus-size-wedding-dresses-catid673003643',
                          'http://www.sophiaprom.com/vintage-wedding-dresses-catid673005542',
                          'http://www.sophiaprom.com/luxury-wedding-dresses-catid673004364',
                          'http://www.sophiaprom.com/wedding-party-dresses-catid673000069',
                          'http://www.sophiaprom.com/bridesmaid-dresses-catid673000136',
                          'http://www.sophiaprom.com/flower-girl-dresses-catid673000137',
                          'http://www.sophiaprom.com/mother-of-the-bride-dresses-catid673000140',
                          'http://www.sophiaprom.com/prom-dresses-2016-catid673004421',]
            feeder.init_test_feeds(start_urls)
        else:
            feeder.init_feeds(spider_name=spider.name,
                feed_type=spider.feed_type)

        # for collection and products relationship
        for feed in feeder.collection_feeds:
            collection = feed[0].strip()
            url = spider.convert_url(feed[2].strip())
            spider.start_urls.append(url)
            collection_set = feeder.collections(url)
            collection_set.add(collection)
            feeder.map_url_collections(url, collection_set)

    def close_spider(self, spider):
        return  # TODO: currently, we don't update crawl time
        if spider.settings['VERSION'] == 'RELEASE':
            feeder = spider.feeder
            feeder.update_next_crawl_datetime()

    def __assert_necessary_attributes(self, item):
        assert_fields = ('title', 'price', 'collections', 'image_urls', 'sku')
        for field in assert_fields:
            if item.get(field) is None:
                raise DropItem("Missing field [%s] in %s" % (field, item))
        if len(item.get('images')) == 0:
            raise DropItem("Download images error.")
    
    def init_to_excel(self , item):
        wb = load_workbook('D:/www/dev-web-crawler/terms-products.xlsx')
        ws = wb.active
        data = []
        data.append(item['referer'])
        data.append(item['url'])
        data.append(item['product_id'])
        ws.append(data)
        wb.save('D:/www/dev-web-crawler/terms-products.xlsx')
        
    
    def store_to_excel(self , item):
        wb = load_workbook('D:/www/dev-web-crawler/products.xlsx')
        ws = wb.active
        data = []
        data.append(item['sku'])
        data.append(item['product_id'])
        data.append('dress')
        data.append(item['title'])
        data.append('')
        data.append(item['price'][len('USD '):])
        data.append(item['list_price'][len('USD '):])
        data.append('1.5')
        data.append('2')
        data.append('Blushing Pink%%Champagne%%Iovry%%White')
        data.append('')
        data.append('')
        data.append('2%%4%%6%%8%%10%%12%%14%%16%%16W%%18W%%20W%%22W%%24W%%26W')
        
        data.append(item['features']['Silhouette'].replace(', ','%%') if 'Silhouette' in item['features'] else '')
        data.append(item['features']['Hemline/Train'].replace(', ','%%') if 'Hemline/Train' in item['features'] else '')
        data.append(item['features']['Neckline'].replace(', ','%%') if 'Neckline' in item['features'] else '')
        data.append('')
        data.append(item['features']['Waist'].replace(', ','%%') if 'Waist' in item['features'] else '')
        data.append(item['features']['Sleeve'].replace(', ','%%') if 'Sleeve' in item['features'] else '')
        data.append(item['features']['Fabric'].replace(', ','%%') if 'Fabric' in item['features'] else '')
        data.append(item['features']['Embellishment'].replace(', ','%%') if 'Embellishment' in item['features'] else '')
        data.append('Yes')
        data.append('Yes')
        data.append(item['features']['Boning'].replace(', ','%%') if 'Boning' in item['features'] else '')
        data.append('')
        data.append(item['features']['Body Shape'].replace(', ','%%') if 'Body Shape' in item['features'] else '')
        data.append(item['features']['Belt Fabric'].replace(', ','%%') if 'Belt Fabric' in item['features'] else '')
        data.append(item['features']['Back Style'].replace(', ','%%') if 'Back Style' in item['features'] else '')
        data.append('')
        data.append(item['features']['Trend'].replace(', ','%%') if 'Trend' in item['features'] else '')
        data.append(item['features']['Style'].replace(', ','%%')  if 'Style' in item['features'] else '')
        data.append(item['features']['Occasion'].replace(', ','%%') if 'Occasion' in item['features'] else '')
        data.append(item['features']['Season'].replace(', ','%%') if 'Season' in item['features'] else '')
        data.append(item['features']['Wedding Venues'].replace(', ','%%') if 'Wedding Venues' in item['features'] else '')
        data.append('5')
        data.append('789')
        
        ws.append(data)
        wb.save('D:/www/dev-web-crawler/products.xlsx')
        
        if len(item['review_list']) != 0:
            wb = load_workbook('D:/www/dev-web-crawler/reviews.xlsx')
            ws = wb.active
            for rev in item['review_list']:
                reviews = []
                reviews.append(item['sku'])
                reviews.append(rev['name'])
                reviews.append(rev['title'])
                reviews.append(rev['content'])
                
                ws.append(reviews)
                wb.save('D:/www/dev-web-crawler/reviews.xlsx')


class AvarshaS3FilesStore(S3FilesStore):
    def __init__(self, *args, **kwargs):
        super(AvarshaS3FilesStore, self).__init__(*args, **kwargs)

    def stat_file(self, path, info):
        def _onsuccess(boto_key):
            checksum = boto_key.etag.strip('"')
            last_modified = boto_key.last_modified
            modified_tuple = rfc822.parsedate_tz(last_modified)
            modified_stamp = int(rfc822.mktime_tz(modified_tuple))
            return {'checksum': checksum,
                    'last_modified': modified_stamp,
                    'width': boto_key.metadata.width,
                    'height': boto_key.metadata.height}
        return self._get_boto_key(path).addCallback(_onsuccess)

class AvarshaImagePipeline(ImagesPipeline):
    def __init__(self, *args, **kwargs):
        self.STORE_SCHEMES['s3'] = AvarshaS3FilesStore
        super(ImagesPipeline, self).__init__(*args, **kwargs)

    def media_to_download(self, request, info):
        def _onsuccess(result):
            if not result:
                return  # returning None force download

            last_modified = result.get('last_modified', None)
            if not last_modified:
                return  # returning None force download

            age_seconds = time.time() - last_modified
            age_days = age_seconds / 60 / 60 / 24
            if age_days > self.EXPIRES:
                return  # returning None force download

            referer = request.headers.get('Referer')
            log.msg(format='File (uptodate): Downloaded %(medianame)s from %(request)s referred in <%(referer)s>',
                    level=log.DEBUG, spider=info.spider,
                    medianame=self.MEDIA_NAME, request=request, referer=referer)
            self.inc_stats(info.spider, 'uptodate')

            checksum = result.get('checksum', None)
            width = result.get('width', None)
            height = result.get('height', None)
            return {'url': request.url, 'path': path, 'checksum': checksum, 'width': width, 'height': height}

        path = self.file_path(request, info=info)
        dfd = defer.maybeDeferred(self.store.stat_file, path, info)
        dfd.addCallbacks(_onsuccess, lambda _: None)
        dfd.addErrback(log.err, self.__class__.__name__ + '.store.stat_file')
        return dfd


    def media_downloaded(self, response, request, info):
        # extend width and height to item['images']
        default_meta = super(ImagesPipeline, self).\
            media_downloaded(response, request, info)
        (width, height) = self.get_image_sizes(response, request)
        return dict(default_meta.items()
            + {'width': width, 'height': height}.items())

    def get_image_sizes(self, response, request):
        orig_image = Image.open(StringIO(response.body))
        width, height = orig_image.size
        return (width, height)
